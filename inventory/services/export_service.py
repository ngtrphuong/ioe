import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ExportService:
    """
    Export service class for exporting data as Excel files
    """
    
    @staticmethod
    def export_to_excel(data, filename, sheet_name='Sheet1'):
        """
        Export data as an Excel file
        :param data: List of dicts, e.g. [{'header1': value1, ...}]
        :param filename: File name
        :param sheet_name: Sheet name
        :return: HttpResponse object
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # Set header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Set cell border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Write headers
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                    cell.border = thin_border
                    # Handle date formatting
                    if isinstance(cell.value, datetime.datetime):
                        cell.number_format = 'yyyy-mm-dd'
                    elif isinstance(cell.value, datetime.date):
                        cell.number_format = 'yyyy-mm-dd'
            
            # Auto fit column width
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                # Base width, plus some for header
                column_width = max(10, len(str(header)) + 2)
                # Check column content for max width
                for row_idx in range(2, len(data) + 2):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        column_width = max(column_width, min(len(str(cell_value)) + 2, 50))  # max 50
                worksheet.column_dimensions[column_letter].width = column_width
        
        # Create response
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def format_member_data_for_export(member_data, start_date, end_date):
        """
        Format member analysis data for export
        """
        # Header info
        header_data = [
            {'Label': 'Report Type', 'Value': 'Member Analysis Report'},
            {'Label': 'Reporting Period', 'Value': f"{start_date} to {end_date}"},
            {'Label': 'Total Members', 'Value': member_data['total_members']},
            {'Label': 'New Members', 'Value': member_data['new_members']},
            {'Label': 'Active Members', 'Value': member_data['active_members']},
            {'Label': 'Member Activity Rate', 'Value': f"{member_data['activity_rate']}%"},
            {'Label': '', 'Value': ''},
        ]
        
        # Member level distribution
        level_data = []
        for level in member_data['level_distribution']:
            level_data.append({
                'Level': level['level__name'],
                'Count': level['count'],
                'Share': f"{(level['count'] / member_data['total_members'] * 100):.1f}%"
            })
        
        # Top members by consumption
        top_members_data = []
        for idx, member in enumerate(member_data['top_members'], 1):
            avg_order = member['period_spend'] / member['period_purchase_count'] if member['period_purchase_count'] > 0 else 0
            top_members_data.append({
                'Rank': idx,
                'Member Name': member['name'],
                'Level': member['level'].name,
                'Phone': member['phone'],
                'Amount Spent': f"VNĐ{member['period_spend']:.2f}",
                'Number of Purchases': member['period_purchase_count'],
                'Avg Order Value': f"VNĐ{avg_order:.2f}"
            })
        
        return {
            'header': header_data,
            'level_distribution': level_data,
            'top_members': top_members_data
        }
    
    @staticmethod
    def export_member_analysis(member_data, start_date, end_date):
        """
        Export member analysis report
        """
        formatted_data = ExportService.format_member_data_for_export(member_data, start_date, end_date)
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Header sheet
        header_sheet = workbook.active
        header_sheet.title = "Overview"
        header_sheet['A1'] = "Member Analysis Report"
        header_sheet['A1'].font = Font(bold=True, size=16)
        header_sheet.merge_cells('A1:B1')
        
        for idx, item in enumerate(formatted_data['header'], 3):
            header_sheet[f'A{idx}'] = item['Label']
            header_sheet[f'B{idx}'] = item['Value']
        
        # Member level distribution sheet
        if formatted_data['level_distribution']:
            level_sheet = workbook.create_sheet(title="Member Level Distribution")
            # Header
            level_sheet['A1'] = "Level"
            level_sheet['B1'] = "Count"
            level_sheet['C1'] = "Share"
            # Header styles
            for cell in level_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            # Data
            for idx, item in enumerate(formatted_data['level_distribution'], 2):
                level_sheet[f'A{idx}'] = item['Level']
                level_sheet[f'B{idx}'] = item['Count']
                level_sheet[f'C{idx}'] = item['Share']
        # Top member ranking sheet
        if formatted_data['top_members']:
            member_sheet = workbook.create_sheet(title="Top Member Ranking")
            headers = ["Rank", "Member Name", "Level", "Phone", "Amount Spent", "Number of Purchases", "Avg Order Value"]
            for col, header in enumerate(headers, 1):
                member_sheet.cell(row=1, column=col, value=header)
                member_sheet.cell(row=1, column=col).font = Font(bold=True)
                member_sheet.cell(row=1, column=col).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            # Data
            for row, item in enumerate(formatted_data['top_members'], 2):
                member_sheet.cell(row=row, column=1, value=item['Rank'])
                member_sheet.cell(row=row, column=2, value=item['Member Name'])
                member_sheet.cell(row=row, column=3, value=item['Level'])
                member_sheet.cell(row=row, column=4, value=item['Phone'])
                member_sheet.cell(row=row, column=5, value=item['Amount Spent'])
                member_sheet.cell(row=row, column=6, value=item['Number of Purchases'])
                member_sheet.cell(row=row, column=7, value=item['Avg Order Value'])
        # Auto column width for each sheet
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
        # Create response
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        # Format date as string for filename
        date_str = datetime.datetime.now().strftime('%Y%m%d')
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Member_Analysis_Report_{date_str}.xlsx"'
        return response 